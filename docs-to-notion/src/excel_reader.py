import openpyxl
from dataclasses import dataclass, field
from typing import List, Optional

@dataclass
class CellData:
    value: str
    row: int
    col: int
    is_bold: bool = False
    is_merged: bool = False
    bg_color: Optional[str] = None
    font_size: Optional[float] = None

@dataclass
class SheetData:
    name: str
    cells: List[List[CellData]] = field(default_factory=list)
    tables: List[dict] = field(default_factory=list)
    headings: List[dict] = field(default_factory=list)
    paragraphs: List[dict] = field(default_factory=list)

def read_excel(file_path: str) -> List[SheetData]:
    """Excelファイルを読み込み、構造化データとして返す"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = []

    for ws in wb.worksheets:
        sheet = SheetData(name=ws.title)
        merged_ranges = list(ws.merged_cells.ranges)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=ws.max_column):
            row_data = []
            for cell in row:
                is_merged = any(cell.coordinate in mr for mr in merged_ranges)
                try:
                    bg = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
                    bg = bg if bg and bg != "00000000" else None
                except Exception:
                    bg = None
                cell_data = CellData(
                    value=str(cell.value) if cell.value is not None else "",
                    row=cell.row,
                    col=cell.column,
                    is_bold=cell.font.bold if cell.font else False,
                    is_merged=is_merged,
                    bg_color=bg,
                    font_size=cell.font.size if cell.font else None,
                )
                row_data.append(cell_data)
            sheet.cells.append(row_data)

        _analyze_structure(sheet)
        sheets.append(sheet)

    return sheets

def _is_row_empty(row_cells: List[CellData]) -> bool:
    return all(c.value == "" for c in row_cells)

def _is_heading_row(row_cells: List[CellData]) -> bool:
    """
    以下の条件を1つ以上満たす場合、見出し行とみなす:
    - 結合セル かつ 太字
    - フォントサイズが12pt以上 かつ 太字
    - 背景色あり かつ 非空
    """
    non_empty = [c for c in row_cells if c.value]
    if not non_empty:
        return False
    first = non_empty[0]
    is_big_bold = first.is_bold and (first.font_size or 0) >= 12
    is_merged_bold = first.is_merged and first.is_bold
    has_bg = any(c.bg_color for c in non_empty)
    return is_big_bold or is_merged_bold or (has_bg and first.is_bold)

def _count_non_empty_cols(row_cells: List[CellData]) -> int:
    return sum(1 for c in row_cells if c.value)

def _analyze_structure(sheet: SheetData):
    """
    シートの構造を解析し、見出し・表・本文に分類する。
    
    アルゴリズム:
    1. 結合セル + 太字 + 大きいフォント → 見出し
    2. 背景色あり + 太字 → セクション見出し
    3. 連続する同列数の行 → テーブル
    4. 空行 → セクション区切り（divider）
    5. 単一セルに長いテキスト → 本文（paragraph）
    """
    elements = []  # {"type": ..., ...}
    
    i = 0
    rows = sheet.cells
    n = len(rows)
    
    while i < n:
        row = rows[i]
        
        # 空行 → divider
        if _is_row_empty(row):
            elements.append({"type": "divider"})
            i += 1
            continue
        
        # 見出し行判定
        if _is_heading_row(row):
            non_empty = [c for c in row if c.value]
            text = " ".join(c.value for c in non_empty)
            # レベル決定: merged+bold+big=1, merged+bold=2, bg+bold=3
            first = non_empty[0]
            if first.is_merged and first.is_bold and (first.font_size or 0) >= 14:
                level = 1
            elif first.is_merged and first.is_bold:
                level = 2
            else:
                level = 3
            elements.append({"type": "heading", "text": text, "level": level})
            i += 1
            continue
        
        # テーブル検出: 次の行も同じ列数の非空行が続く
        col_count = _count_non_empty_cols(row)
        if col_count > 1:
            # テーブル候補: 連続する行をまとめる
            table_rows = []
            while i < n and not _is_row_empty(rows[i]) and not _is_heading_row(rows[i]):
                table_rows.append([c.value for c in rows[i]])
                i += 1
            if len(table_rows) >= 1:
                headers = table_rows[0]
                data = table_rows[1:]
                elements.append({"type": "table", "headers": headers, "rows": data})
            continue
        
        # 単一セル → 本文
        text = " ".join(c.value for c in row if c.value)
        if len(text) > 100:
            elements.append({"type": "paragraph", "text": text})
        elif text:
            elements.append({"type": "paragraph", "text": text})
        i += 1
    
    # SheetDataにelementsを格納（markdown_converterで使う）
    sheet._elements = elements
    # tablesにも互換性のために追加
    for el in elements:
        if el["type"] == "table":
            sheet.tables.append(el)


def _iterate_elements(sheet: SheetData):
    """_analyze_structureが作ったelementsを返す（markdown_converter用）"""
    return getattr(sheet, "_elements", sheet.tables)
